"""
TV Top 100 MCap Fundamental Screener — xuất Excel + CAN SLIM Score
===================================================================
pip install tradingview-screener pandas openpyxl yfinance
python tv_top100_screener.py
python tv_top100_screener.py --market nasdaq --top 200
"""

import argparse, sys, time, warnings
from datetime import datetime
import numpy as np
warnings.filterwarnings("ignore")

try:
    from tradingview_screener import Query, Column
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
except ImportError as e:
    print(f"❌ Thiếu thư viện: {e}\n   pip install tradingview-screener pandas openpyxl yfinance")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════
# CAN SLIM CRITERIA
# ═══════════════════════════════════════════════════════════════
CANSLIM = {
    "C":   dict(label="C — EPS Qtr",   field="EPS Qtr%",      thr=25,  op="gt", desc="EPS Quarterly YoY > 25%"),
    "A":   dict(label="A — EPS Ann",   field="EPS Annual%",   thr=20,  op="gt", desc="EPS Annual YoY > 20%"),
    "S":   dict(label="S — Sales",     field="Rev Qtr%",      thr=20,  op="gt", desc="Revenue Quarterly YoY > 20%"),
    "L":   dict(label="L — RS (vs Idx)", field="RS_1Y%",        thr=0,   op="gt", desc="1Y outperform market index"),
    "Q":   dict(label="Quality GM",    field="Gross Margin%", thr=40,  op="gt", desc="Gross Margin > 40%"),
    "R":   dict(label="ROE",           field="ROE%",          thr=17,  op="gt", desc="ROE > 17%"),
    "M":   dict(label="M — EPS Accel",  field="EPS Qtr%",      thr=0,   op="accel", desc="EPS Quarterly YoY > Annual YoY AND > 0 (accelerating earnings momentum)"),
    "D":   dict(label="Debt OK",       field="D/E",           thr=1.5, op="lt", desc="D/E < 1.5"),
    "N":   dict(label="N — 52W High",  field="52W_High%",     thr=90,  op="gt", desc="Price ≥ 90% of 52-Week High"),
    "MKT": dict(label="Market",        field="Market_OK",     thr=0.5, op="gt", desc="Market index above MA50 & MA200"),
}
CS_KEYS = ["C", "A", "S", "L", "Q", "R", "M", "D", "N", "MKT"]
N_CS    = len(CS_KEYS)

FETCH_COLS = [
    "name", "description", "close", "volume", "market_cap_basic",
    "sector",
    "earnings_per_share_diluted_yoy_growth_fy",
    "earnings_per_share_diluted_yoy_growth_fq",
    "total_revenue_yoy_growth_fy",
    "total_revenue_yoy_growth_fq",
    "gross_margin", "net_margin", "return_on_equity", "debt_to_equity",
    "price_earnings_ttm", "price_sales_current",
    "relative_volume_10d_calc",
    "Perf.W", "Perf.1M", "Perf.3M", "Perf.6M", "Perf.Y",
    # Quality Compounder fields
    "return_on_invested_capital", "operating_margin",
    "enterprise_value_ebitda_ttm", "free_cash_flow_per_share_ttm",
    "current_ratio",
]

RENAME = {
    "name": "Ticker", "description": "Tên Công Ty", "close": "Price ($)", "market_cap_basic": "MCap ($B)",
    "sector": "Sector",
    "earnings_per_share_diluted_yoy_growth_fy": "EPS Annual%",
    "earnings_per_share_diluted_yoy_growth_fq": "EPS Qtr%",
    "total_revenue_yoy_growth_fy": "Rev Annual%",
    "total_revenue_yoy_growth_fq": "Rev Qtr%",
    "gross_margin": "Gross Margin%", "net_margin": "Net Margin%",
    "return_on_equity": "ROE%", "debt_to_equity": "D/E",
    "price_earnings_ttm": "P/E", "price_sales_current": "P/S",
    "relative_volume_10d_calc": "Rel Vol",
    "Perf.W": "1W%", "Perf.1M": "1M%", "Perf.3M": "3M%",
    "Perf.6M": "6M%", "Perf.Y": "1Y%",
    "return_on_invested_capital": "ROIC%",
    "operating_margin": "Op Margin%",
    "enterprise_value_ebitda_ttm": "EV/EBITDA",
    "free_cash_flow_per_share_ttm": "FCF/sh",
    "current_ratio": "Current Ratio",
}

DATA_HEADERS = [
    "Ticker", "Tên Công Ty", "Sector", "Price ($)", "MCap ($B)",
    "Moat Proxy", "Moat Score",
    "EPS Annual%", "EPS Qtr%", "Rev Annual%", "Rev Qtr%",
    "Gross Margin%", "Net Margin%", "ROE%", "D/E",
    "P/E", "P/S", "Rel Vol",
    "1W%", "1M%", "3M%", "6M%", "1Y%",
]
# 1=Ticker,2=TênCT,3=Sector,4=Price,5=MCap
# 6=MoatProxy,7=MoatScore
# 8=EPS_A,9=EPS_Q,10=Rev_A,11=Rev_Q
# 12=GM,13=NM,14=ROE,15=DE,16=PE,17=PS,18=RVol
# 19=1W,20=1M,21=3M,22=6M,23=1Y
PCT_COLS = {8,9,10,11,12,13,14,19,20,21,22,23}


# ═══════════════════════════════════════════════════════════════
# MOAT — lấy từ yfinance (5yr ROE avg + 5yr GM avg)
# Logic giống moat_score() trong screen_accumulate.py
# ═══════════════════════════════════════════════════════════════
MOAT_PROXY_MAP = {
    "Technology":            "Switching Cost / Network Effect",
    "Healthcare":            "Intangible Assets (Patents)",
    "Consumer Defensive":    "Brand / Distribution Network",
    "Financial Services":    "Network Effect / Scale",
    "Industrials":           "Cost Advantage / Scale",
    "Communication Services":"Network Effect",
    "Consumer Cyclical":     "Brand / Cost Advantage",
    "Basic Materials":       "Cost Advantage / Efficient Scale",
    "Energy":                "Cost Advantage",
    "Utilities":             "Efficient Scale (Regulated)",
    "Real Estate":           "Efficient Scale / Location",
}

MOAT_SCORE_STYLE = {
    "WIDE  ★★★": ("1A5C2B", "C6EFCE"),
    "NARROW ★★":  ("1A3A5C", "DDEEFF"),
    "UNCERTAIN ★": ("7D6608", "FFF2CC"),
    "WEAK":        ("9C0006", "FFC7CE"),
}

QC_SIGNAL_STYLE = {
    "🏆 COMPOUNDER": ("1A5C2B", "C6EFCE"),
    "⭐ QUALITY":     ("1A3A5C", "DDEEFF"),
    "○ AVERAGE":     ("7D6608", "FFF2CC"),
    "✗ WEAK":        ("9C0006", "FFC7CE"),
}

EQ_BADGE_STYLE = {
    "💚 Cash Backed":   ("1A5C2B", "C6EFCE"),
    "🟡 Mixed":         ("7D6608", "FFF2CC"),
    "🔴 Accrual Heavy": ("9C0006", "FFC7CE"),
}

def _sv(s, i=0):
    try:
        v = s.iloc[i] if hasattr(s, "iloc") else s
        return float(v) if pd.notna(v) else None
    except: return None

def _get_row(df_fin, *keys):
    for k in keys:
        if k in df_fin.index:
            s = df_fin.loc[k]
            return s if hasattr(s, "iloc") else pd.Series([s])
    return pd.Series(dtype=float)

def _safe_mean(lst):
    v = [x for x in lst if x is not None]
    return round(float(np.mean(v)), 1) if v else None

def _moat_score_from_metrics(roe_avg, gm_avg, sector=""):
    """Copy từ moat_score() trong screen_accumulate.py"""
    r = roe_avg or 0;  g = gm_avg or 0
    if sector in ("Utilities", "Real Estate"):
        return "NARROW ★★" if r >= 10 and g >= 20 else "UNCERTAIN ★"
    if   r >= 20 and g >= 50: return "WIDE  ★★★"
    elif r >= 15 and g >= 35: return "NARROW ★★"
    elif r >= 10:             return "UNCERTAIN ★"
    return "WEAK"

def fetch_moat_yfinance(ticker: str, sector: str = "") -> tuple:
    """
    Lấy 5yr ROE avg + 5yr GM avg từ yfinance.
    Trả về (moat_proxy, moat_score).
    Giống fetch() trong screen_accumulate.py.
    """
    try:
        import yfinance as yf, sys, io
        tk   = yf.Ticker(ticker)
        _stderr, sys.stderr = sys.stderr, io.StringIO()
        try:
            info = tk.info or {}
        finally:
            sys.stderr = _stderr

        # Sector từ yfinance (chính xác hơn TV)
        yf_sector = info.get("sector", "") or sector
        proxy = MOAT_PROXY_MAP.get(yf_sector, MOAT_PROXY_MAP.get(sector, "Cost Advantage"))

        # TTM fallback
        roe_ttm = (info.get("returnOnEquity") or 0) * 100 or None
        gm_ttm  = (info.get("grossMargins")   or 0) * 100 or None

        # 5yr ROE từ income + balance sheet
        roe_avg = None
        try:
            inc = tk.income_stmt.sort_index(axis=1, ascending=False)
            bal = tk.balance_sheet.sort_index(axis=1, ascending=False)
            ni  = _get_row(inc, "Net Income", "NetIncome")
            eq  = _get_row(bal, "Stockholders Equity", "StockholdersEquity",
                           "Total Stockholders Equity", "Common Stock Equity")
            roe_list = []
            for i in range(min(5, min(len(ni), len(eq)))):
                n = _sv(ni, i); e = _sv(eq, i)
                roe_list.append(round(n/e*100,1) if n and e and e != 0 else None)
            roe_avg = _safe_mean(roe_list)
        except: pass

        # 5yr GM từ income stmt
        gm_avg = None
        try:
            inc2 = tk.income_stmt.sort_index(axis=1, ascending=False)
            gp   = _get_row(inc2, "Gross Profit", "GrossProfit")
            rev  = _get_row(inc2, "Total Revenue", "TotalRevenue")
            gm_list = []
            for i in range(min(5, min(len(gp), len(rev)))):
                g = _sv(gp, i); r = _sv(rev, i)
                gm_list.append(round(g/r*100,1) if g and r and r != 0 else None)
            gm_avg = _safe_mean(gm_list)
        except: pass

        # Dùng 5yr avg nếu có, fallback TTM
        score = _moat_score_from_metrics(
            roe_avg or roe_ttm,
            gm_avg  or gm_ttm,
            yf_sector
        )

        # 52-week high proximity
        w52_high  = info.get("fiftyTwoWeekHigh")
        cur_price = (info.get("currentPrice") or info.get("regularMarketPrice")
                     or info.get("previousClose"))
        w52_pct = (round(cur_price / w52_high * 100, 1)
                   if w52_high and cur_price and w52_high > 0 else None)

        # Earnings quality — FCF / Net Income ratio (TTM)
        eq_badge = None
        try:
            cf   = tk.cashflow.sort_index(axis=1, ascending=False)
            inc3 = tk.income_stmt.sort_index(axis=1, ascending=False)
            fcf_row = _get_row(cf,   "Free Cash Flow",     "FreeCashFlow")
            ni_row  = _get_row(inc3, "Net Income",          "NetIncome")
            if fcf_row.empty:
                ocf = _get_row(cf, "Operating Cash Flow",  "OperatingCashFlow")
                cex = _get_row(cf, "Capital Expenditure",  "CapitalExpenditure")
                ocf_v = _sv(ocf, 0); cex_v = _sv(cex, 0)
                fcf_v = (ocf_v + (cex_v or 0)) if ocf_v is not None else None
            else:
                fcf_v = _sv(fcf_row, 0)
            ni_v = _sv(ni_row, 0)
            if fcf_v is not None and ni_v is not None and ni_v != 0:
                ratio = fcf_v / ni_v
                eq_badge = ("💚 Cash Backed"  if ratio >= 0.8
                            else "🟡 Mixed"       if ratio >= 0.3
                            else "🔴 Accrual Heavy")
        except:
            pass

        return proxy, score, w52_pct, eq_badge

    except Exception as e:
        # Fallback: sector hardcode + UNCERTAIN
        proxy = MOAT_PROXY_MAP.get(sector, "Cost Advantage")
        return proxy, "UNCERTAIN ★", None, None


def build_moat_cache(tickers: list, sectors: dict) -> dict:
    """
    Fetch moat cho toàn bộ danh sách tickers từ yfinance.
    Trả về dict: {ticker: (proxy, score)}
    """
    cache = {}
    total = len(tickers)
    print(f"\n  🏰 Fetch Moat từ yfinance ({total} tickers)...")
    for i, ticker in enumerate(tickers, 1):
        sector = sectors.get(ticker, "")
        print(f"  [{i:>3}/{total}] {ticker:<8}", end="", flush=True)
        proxy, score, w52_pct, eq_badge = fetch_moat_yfinance(ticker, sector)
        cache[ticker] = (proxy, score, w52_pct, eq_badge)
        eq_str = f" [{eq_badge}]" if eq_badge else ""
        print(f" {score}{eq_str}")
        time.sleep(0.4)   # tránh rate limit
    print()
    return cache


_MKT_INDEX = {
    "america":   ("america",   "SP:SPX"),
    "nasdaq":    ("america",   "NASDAQ:NDX"),
    "nyse":      ("america",   "NYSE:NYA"),
    "euronext":  ("global",    "TVC:CAC40"),
    "hong_kong": ("hongkong",  "HSI:HSI"),
    "vietnam":   ("vietnam",   "HOSE:VNINDEX"),
}

def fetch_market_direction(market="america"):
    """Returns (market_ok, index_1y).
    market_ok: True nếu index trên MA50 & MA200; index_1y: % 1-year perf của index."""
    import requests
    tv_market, ticker = _MKT_INDEX.get(market.lower(), ("america", "SP:SPX"))
    try:
        r = requests.post(
            f"https://scanner.tradingview.com/{tv_market}/scan",
            json={"symbols": {"tickers": [ticker], "query": {"types": []}},
                  "columns": ["close", "SMA50", "SMA200", "Perf.Y"]},
            timeout=10)
        d = r.json()["data"][0]["d"]
        price, ma50, ma200, perf_y = d[0], d[1], d[2], d[3]
        if price is None or ma50 is None or ma200 is None:
            return None, None
        market_ok = float(price) > float(ma50) and float(price) > float(ma200)
        index_1y = float(perf_y) if perf_y is not None else None
        return market_ok, index_1y
    except Exception:
        return None, None


# ─── Colors ─────────────────────────────────────────────────────────────────
C_NAVY="1B3A5C"; C_DARK="0D2137"; C_WHITE="FFFFFF"; C_ALT="F0F4FA"; C_BRD="C5D5E8"
CG_BG="C6EFCE"; CG_FG="276221"
CR_BG="FFC7CE"; CR_FG="9C0006"
CY_BG="FFEB9C"; CY_FG="9C6500"
CB_BG="DDEEFF"; CB_FG="1B3A5C"

SIGNAL_STYLE = {
    "🟢 STRONG BUY": (CG_FG, CG_BG),
    "🔵 BUY":        (CB_FG, CB_BG),
    "🟡 WATCH":      (CY_FG, CY_BG),
    "🔴 SKIP":       (CR_FG, CR_BG),
}

def F(bold=False, size=9, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)
def BG(c): return PatternFill("solid", fgColor=c)
def AL(h="center", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def BD(c=C_BRD):
    s = Side(style="thin", color=c); return Border(left=s, right=s, top=s, bottom=s)
def CL(i): return get_column_letter(i)


# ═══════════════════════════════════════════════════════════════
# SCORING
# ═══════════════════════════════════════════════════════════════
def score_canslim(df, moat_cache: dict = None, market_ok=None, index_1y=None):
    # Populate 52W_High% from moat_cache (3rd element of tuple)
    if moat_cache:
        def _get_w52(ticker):
            t = moat_cache.get(ticker)
            return t[2] if t and len(t) >= 3 else None
        df["52W_High%"] = df["Ticker"].apply(_get_w52)
    else:
        df["52W_High%"] = None

    # Relative Strength vs index for L criterion
    if index_1y is not None and "1Y%" in df.columns:
        df["RS_1Y%"] = df["1Y%"] - index_1y
    else:
        df["RS_1Y%"] = df["1Y%"] if "1Y%" in df.columns else None

    # Market direction — same value for all rows
    df["Market_OK"] = market_ok

    def chk(row, key):
        cfg = CANSLIM[key]
        if key == "D" and row.get("Sector", "") in ("Financial Services", "Finance"):
            return None  # D/E không áp dụng cho ngành tài chính (banks/insurance dùng leverage cấu trúc)
        if key == "M":
            qtr = row.get("EPS Qtr%"); ann = row.get("EPS Annual%")
            if qtr is None or (isinstance(qtr, float) and pd.isna(qtr)): return None
            if ann is None or (isinstance(ann, float) and pd.isna(ann)): return None
            return float(qtr) > float(ann) and float(qtr) > 0
        v = row.get(cfg["field"])
        if v is None or (isinstance(v, float) and pd.isna(v)): return None
        return v > cfg["thr"] if cfg["op"] == "gt" else v < cfg["thr"]

    for key in CS_KEYS:
        df[f"CS_{key}"] = df.apply(lambda r: chk(r, key), axis=1)

    df["CS_Score"] = df[[f"CS_{k}" for k in CS_KEYS]].apply(
        lambda r: sum(1 for v in r if v is True), axis=1)

    def sig(s):
        if s >= round(N_CS * 0.80): return "🟢 STRONG BUY"
        if s >= round(N_CS * 0.625): return "🔵 BUY"
        if s >= round(N_CS * 0.375): return "🟡 WATCH"
        return "🔴 SKIP"
    df["CS_Signal"] = df["CS_Score"].apply(sig)

    # ── MOAT: dùng yfinance cache nếu có, fallback sector ──────
    def apply_moat(row):
        ticker = row.get("Ticker", "")
        sector = row.get("Sector", "")
        if moat_cache and ticker in moat_cache:
            t = moat_cache[ticker]
            return t[0], t[1]   # proxy, score (bỏ qua w52_pct)
        # Fallback: tính từ TV TTM data (giống cũ)
        gm  = row.get("Gross Margin%"); gm  = gm  if (gm  is not None and not (isinstance(gm,  float) and pd.isna(gm)))  else None
        roe = row.get("ROE%");          roe = roe if (roe is not None and not (isinstance(roe, float) and pd.isna(roe))) else None
        proxy = MOAT_PROXY_MAP.get(sector, "Cost Advantage")
        score = _moat_score_from_metrics(roe, gm, sector)
        return proxy, score

    moat = df.apply(apply_moat, axis=1)
    df["Moat Proxy"] = moat.apply(lambda x: x[0])
    df["Moat Score"] = moat.apply(lambda x: x[1])

    # Earnings quality badge (only when moat_cache / yfinance is available)
    if moat_cache:
        def _get_eq(ticker):
            t = moat_cache.get(ticker)
            return t[3] if t and len(t) >= 4 else None
        df["EQ_Badge"] = df["Ticker"].apply(_get_eq)
    else:
        df["EQ_Badge"] = None

    _MW = {"WIDE  ★★★": 1.2, "NARROW ★★": 1.1, "UNCERTAIN ★": 1.0, "WEAK": 0.85}
    df["Conviction"] = df.apply(
        lambda r: round(r["CS_Score"] * _MW.get(r.get("Moat Score", ""), 1.0), 1),
        axis=1)
    return df


# ═══════════════════════════════════════════════════════════════
# FETCH
# ═══════════════════════════════════════════════════════════════
def fetch_data(market, top):
    print(f"  ⏳ Query TradingView ({market.upper()}, top {top} MCap)...", end="", flush=True)
    try:
        _, df = (Query().set_markets(market.lower()).select(*FETCH_COLS)
                 .order_by("market_cap_basic", ascending=False).limit(top).get_scanner_data())
        print(f" ✅ {len(df)} mã"); return df
    except Exception as e:
        print(f"\n❌ {e}"); sys.exit(1)

def clean_df(df):
    df = df.rename(columns=RENAME)
    df["MCap ($B)"] = df["MCap ($B)"] / 1e9
    return df


# ═══════════════════════════════════════════════════════════════
# EXCEL — giữ nguyên như cũ
# ═══════════════════════════════════════════════════════════════
def write_excel(df, path, market, top, use_yf=False, progress_cb=None):
    def _pb(pct, msg=""):
        if progress_cb:
            progress_cb(pct, msg)

    wb = Workbook()
    _pb(5,  "Building main sheet…")
    _main_sheet(wb, df, market, top)

    if "QC_Score" in df.columns:
        _pb(20, "QC sheet…")
        _qc_sheet(wb, df)

    _US_MARKETS = {"america", "nasdaq", "nyse"}
    can_yf = use_yf and market.lower() in _US_MARKETS
    if can_yf:
        n = len(df)
        def _fin_cb(idx):
            _pb(25 + int(idx / max(n, 1) * 45),
                f"Financials  {idx}/{n}…")
        _financials_sheet(wb, df, progress_cb=_fin_cb)
    else:
        if use_yf and market.lower() not in _US_MARKETS:
            print(f"  ⚠  yfinance không hỗ trợ {market.upper()} — bỏ qua sheet Financials")
        _pb(25, "")

    _pb(75, "Dashboard sheet…")
    _dashboard_sheet(wb, df)
    _pb(95, "Saving file…")
    wb.save(path)
    _pb(100, "")
    print(f"  💾 {path}")


def _financials_sheet(wb, df, progress_cb=None):
    import yfinance as yf
    from openpyxl.chart import BarChart, Reference

    ws = wb.create_sheet("Financials")
    ws.sheet_view.showGridLines = False
    ws.sheet_tab_color = "0E6655"

    N_QTR, N_YR = 8, 4
    C_LBL  = 1
    C_QTR0 = 2
    C_SEP  = C_QTR0 + N_QTR      # col 10
    C_YR0  = C_SEP + 1            # col 11
    LAST_C = C_YR0 + N_YR - 1    # col 14
    C_GAP  = LAST_C + 1           # col 15 — gap before chart area
    C_CH1  = C_GAP + 1            # col 16 — Revenue quarterly
    C_CH2  = C_CH1 + 7            # col 23 — Revenue annual
    C_CH3  = C_CH2 + 7            # col 30 — EPS quarterly
    C_CH4  = C_CH3 + 7            # col 37 — EPS annual
    C_CH5  = C_CH4 + 7            # col 44 — ROE quarterly
    C_CH6  = C_CH5 + 7            # col 51 — ROE annual
    C_CH7  = C_CH6 + 7            # col 58 — CF/Share quarterly
    C_CH8  = C_CH7 + 7            # col 65 — CF/Share annual
    CHART_SPACER = 6
    CHART_H, CHART_W = 4.5, 9.0

    ws.column_dimensions[CL(C_LBL)].width = 15
    for ci in range(C_QTR0, C_SEP):
        ws.column_dimensions[CL(ci)].width = 10
    ws.column_dimensions[CL(C_SEP)].width = 2
    for ci in range(C_YR0, LAST_C + 1):
        ws.column_dimensions[CL(ci)].width = 11
    ws.column_dimensions[CL(C_GAP)].width = 2

    # Title
    ws.merge_cells(f"A1:{CL(LAST_C)}1")
    c = ws.cell(1, 1, "FINANCIAL HISTORY  ·  REVENUE & EPS DILUTED  (nguồn: yfinance)")
    c.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    c.fill = BG("0D2137"); c.alignment = AL()
    ws.row_dimensions[1].height = 30

    # Sub-header: quarters vs years
    ws.merge_cells(f"{CL(C_QTR0)}2:{CL(C_SEP-1)}2")
    c2 = ws.cell(2, C_QTR0, f"◀  {N_QTR} QUÝ GẦN NHẤT  ▶")
    c2.font = F(True, 9, "FFFFFF"); c2.fill = BG("1B4F72"); c2.alignment = AL()
    ws.merge_cells(f"{CL(C_YR0)}2:{CL(LAST_C)}2")
    c3 = ws.cell(2, C_YR0, f"◀  {N_YR} NĂM GẦN NHẤT  ▶")
    c3.font = F(True, 9, "FFFFFF"); c3.fill = BG("0E6655"); c3.alignment = AL()
    ws.cell(2, C_LBL).fill = BG("0D2137")
    ws.cell(2, C_SEP).fill = BG("F0F4FA")
    ws.row_dimensions[2].height = 16

    cur = 3
    tickers = df["Ticker"].tolist()
    total   = len(tickers)

    for idx, ticker in enumerate(tickers, 1):
        if progress_cb: progress_cb(idx)
        print(f"  📊 Financials {idx}/{total}: {ticker}", end="\r", flush=True)
        name = str(df.loc[df["Ticker"] == ticker, "Tên Công Ty"].values[0]
                   if len(df.loc[df["Ticker"] == ticker]) else "")

        try:
            yft  = yf.Ticker(ticker)
            qfin = yft.quarterly_financials
            afin = yft.financials
            qbal = yft.quarterly_balance_sheet
            abal = yft.balance_sheet

            qtr_rev = (qfin.loc["Total Revenue"].sort_index().tail(N_QTR) / 1e9
                       if "Total Revenue" in qfin.index else pd.Series(dtype=float))
            qtr_eps = (qfin.loc["Diluted EPS"].sort_index().tail(N_QTR)
                       if "Diluted EPS" in qfin.index else pd.Series(dtype=float))
            ann_rev = (afin.loc["Total Revenue"].sort_index().dropna().tail(N_YR) / 1e9
                       if "Total Revenue" in afin.index else pd.Series(dtype=float))
            ann_eps = (afin.loc["Diluted EPS"].sort_index().dropna().tail(N_YR)
                       if "Diluted EPS" in afin.index else pd.Series(dtype=float))

            def _row(bs, *keys):
                for k in keys:
                    if k in bs.index: return bs.loc[k].sort_index()
                return pd.Series(dtype=float)

            qtr_ni  = _row(qfin, "Net Income", "Net Income Common Stockholders")
            ann_ni  = _row(afin, "Net Income", "Net Income Common Stockholders")
            qtr_eq  = _row(qbal, "Stockholders Equity", "Common Stock Equity",
                           "Total Equity Gross Minority Interest")
            ann_eq  = _row(abal, "Stockholders Equity", "Common Stock Equity",
                           "Total Equity Gross Minority Interest")

            def _roe_map(ni_s, eq_s, annualize=False):
                ni_sorted = ni_s.sort_index()
                eq_sorted = eq_s.sort_index()
                # TTM rolling: sum 4 quarters instead of multiplying 1 quarter × 4
                ni_use = ni_sorted.rolling(4, min_periods=1).sum() if annualize else ni_sorted
                common = ni_use.index.intersection(eq_sorted.index)
                out = {}
                for dt in common:
                    ni = float(ni_use.loc[dt]); eq = float(eq_sorted.loc[dt])
                    if pd.notna(ni) and pd.notna(eq) and eq != 0:
                        out[dt] = ni / abs(eq)
                return out

            qtr_roe = _roe_map(qtr_ni, qtr_eq.tail(N_QTR), annualize=True)  # full qtr_ni so rolling has enough history
            ann_roe = _roe_map(ann_ni.dropna().tail(N_YR), ann_eq.dropna().tail(N_YR))

            qcf = yft.quarterly_cashflow
            acf = yft.cashflow
            qtr_ocf = _row(qcf, "Operating Cash Flow",
                           "Cash Flow From Continuing Operating Activities")
            ann_ocf = _row(acf, "Operating Cash Flow",
                           "Cash Flow From Continuing Operating Activities")
            qtr_sh  = _row(qbal, "Ordinary Shares Number", "Share Issued",
                           "Common Stock Shares Outstanding")
            ann_sh  = _row(abal, "Ordinary Shares Number", "Share Issued",
                           "Common Stock Shares Outstanding")

            def _cfps_map(ocf_s, sh_s, tail_n):
                ocf_t = ocf_s.sort_index().tail(tail_n)
                common = ocf_t.index.intersection(sh_s.index)
                out = {}
                for dt in common:
                    ocf = float(ocf_t.loc[dt]); sh = float(sh_s.loc[dt])
                    if pd.notna(ocf) and pd.notna(sh) and sh != 0:
                        out[dt] = ocf / sh
                return out

            qtr_cfps = _cfps_map(qtr_ocf, qtr_sh, N_QTR)
            ann_cfps = _cfps_map(ann_ocf.dropna(), ann_sh, N_YR)
        except Exception:
            continue

        stock_row = cur

        # ── Stock header ──────────────────────────────────────────
        ws.merge_cells(f"A{cur}:{CL(LAST_C)}{cur}")
        c = ws.cell(cur, 1, f"  {ticker}  —  {name}")
        c.font = F(True, 10, "FFFFFF"); c.fill = BG("1B3A5C"); c.alignment = AL("left")
        ws.row_dimensions[cur].height = 18; cur += 1

        # ── Column headers: dates ─────────────────────────────────
        col_hdr_row = cur
        ws.cell(cur, C_LBL, "Chỉ số").font      = F(True, 8, "FFFFFF")
        ws.cell(cur, C_LBL).fill                 = BG("2C4F7C")
        ws.cell(cur, C_LBL).alignment            = AL()
        ws.cell(cur, C_LBL).border               = BD()
        ws.cell(cur, C_SEP).fill                 = BG("F0F4FA")

        qtr_idx = qtr_rev.index if len(qtr_rev) else qtr_eps.index
        for i, dt in enumerate(qtr_idx):
            lbl = dt.strftime("%b'%y") if hasattr(dt, "strftime") else str(dt)[:7]
            c = ws.cell(cur, C_QTR0 + i, lbl)
            c.font = F(True, 8, "FFFFFF"); c.fill = BG("1B4F72")
            c.alignment = AL(); c.border = BD()

        yr_idx = ann_rev.index if len(ann_rev) else ann_eps.index
        for i, dt in enumerate(yr_idx):
            lbl = f"FY{dt.year}" if hasattr(dt, "year") else str(dt)[:4]
            c = ws.cell(cur, C_YR0 + i, lbl)
            c.font = F(True, 8, "FFFFFF"); c.fill = BG("0E6655")
            c.alignment = AL(); c.border = BD()

        ws.row_dimensions[cur].height = 16; cur += 1

        # ── Revenue row ───────────────────────────────────────────
        rev_row = cur
        c = ws.cell(cur, C_LBL, "Revenue ($B)")
        c.font = F(True, 9, "1B4F72"); c.fill = BG("EBF5FB")
        c.alignment = AL("left"); c.border = BD()
        ws.cell(cur, C_SEP).fill = BG("F0F4FA")

        for i, v in enumerate(qtr_rev.values):
            cell = ws.cell(cur, C_QTR0 + i, round(float(v), 2) if pd.notna(v) else None)
            cell.font = F(size=9, color="1B4F72"); cell.fill = BG("EBF5FB")
            cell.alignment = AL(); cell.border = BD()
            if pd.notna(v): cell.number_format = "#,##0.00"
        for i, v in enumerate(ann_rev.values):
            cell = ws.cell(cur, C_YR0 + i, round(float(v), 2) if pd.notna(v) else None)
            cell.font = F(True, 9, "1B4F72"); cell.fill = BG("DDEEFF")
            cell.alignment = AL(); cell.border = BD()
            if pd.notna(v): cell.number_format = "#,##0.00"

        ws.row_dimensions[cur].height = 17; cur += 1

        # ── EPS row ───────────────────────────────────────────────
        eps_row = cur
        c = ws.cell(cur, C_LBL, "EPS Diluted ($)")
        c.font = F(True, 9, "276221"); c.fill = BG("E8F8F5")
        c.alignment = AL("left"); c.border = BD()
        ws.cell(cur, C_SEP).fill = BG("F0F4FA")

        for i, v in enumerate(qtr_eps.values):
            fv = float(v) if pd.notna(v) else None
            cell = ws.cell(cur, C_QTR0 + i, fv)
            cell.font = F(size=9, color="276221" if (fv or 0) >= 0 else "9C0006")
            cell.fill = BG("E8F8F5"); cell.alignment = AL(); cell.border = BD()
            if fv is not None: cell.number_format = '$#,##0.00'
        for i, v in enumerate(ann_eps.values):
            fv = float(v) if pd.notna(v) else None
            cell = ws.cell(cur, C_YR0 + i, fv)
            cell.font = F(True, 9, "276221" if (fv or 0) >= 0 else "9C0006")
            cell.fill = BG("C6EFCE"); cell.alignment = AL(); cell.border = BD()
            if fv is not None: cell.number_format = '$#,##0.00'

        ws.row_dimensions[cur].height = 17; cur += 1

        # ── ROE row ───────────────────────────────────────────────
        roe_row = cur
        c = ws.cell(cur, C_LBL, "ROE (annualized)")
        c.font = F(True, 9, "4A235A"); c.fill = BG("F4ECF7")
        c.alignment = AL("left"); c.border = BD()
        ws.cell(cur, C_SEP).fill = BG("F0F4FA")

        qtr_idx = qtr_rev.index if len(qtr_rev) else qtr_eps.index
        for i, dt in enumerate(qtr_idx):
            if dt in qtr_roe:
                v = qtr_roe[dt]
                cell = ws.cell(cur, C_QTR0 + i, round(v, 4))
                cell.font = F(size=9, color="4A235A" if v >= 0 else "9C0006")
                cell.fill = BG("F4ECF7"); cell.alignment = AL(); cell.border = BD()
                cell.number_format = '0.0%'

        yr_idx = ann_rev.index if len(ann_rev) else ann_eps.index
        for i, dt in enumerate(yr_idx):
            if dt in ann_roe:
                v = ann_roe[dt]
                cell = ws.cell(cur, C_YR0 + i, round(v, 4))
                cell.font = F(True, 9, "4A235A" if v >= 0 else "9C0006")
                cell.fill = BG("E8DAEF"); cell.alignment = AL(); cell.border = BD()
                cell.number_format = '0.0%'

        ws.row_dimensions[cur].height = 17; cur += 1

        # ── CF/Share row ──────────────────────────────────────────
        cfps_row = cur
        c = ws.cell(cur, C_LBL, "CF/Share ($)")
        c.font = F(True, 9, "0E6655"); c.fill = BG("E8F6F3")
        c.alignment = AL("left"); c.border = BD()
        ws.cell(cur, C_SEP).fill = BG("F0F4FA")

        qtr_idx = qtr_rev.index if len(qtr_rev) else qtr_eps.index
        for i, dt in enumerate(qtr_idx):
            if dt in qtr_cfps:
                v = qtr_cfps[dt]
                cell = ws.cell(cur, C_QTR0 + i, round(v, 2))
                cell.font = F(size=9, color="0E6655" if v >= 0 else "9C0006")
                cell.fill = BG("E8F6F3"); cell.alignment = AL(); cell.border = BD()
                cell.number_format = '$#,##0.00'

        yr_idx = ann_rev.index if len(ann_rev) else ann_eps.index
        for i, dt in enumerate(yr_idx):
            if dt in ann_cfps:
                v = ann_cfps[dt]
                cell = ws.cell(cur, C_YR0 + i, round(v, 2))
                cell.font = F(True, 9, "0E6655" if v >= 0 else "9C0006")
                cell.fill = BG("D0EDE8"); cell.alignment = AL(); cell.border = BD()
                cell.number_format = '$#,##0.00'

        ws.row_dimensions[cur].height = 17; cur += 1

        # ── Revenue QoQ % row ─────────────────────────────────────
        c = ws.cell(cur, C_LBL, "Rev QoQ %")
        c.font = F(True, 8, "155680"); c.fill = BG("D6EEF8")
        c.alignment = AL("left"); c.border = BD()
        ws.cell(cur, C_SEP).fill = BG("F0F4FA")

        rv = list(qtr_rev.values)
        for i, v in enumerate(rv):
            if i == 0 or not pd.notna(v) or not pd.notna(rv[i-1]) or rv[i-1] == 0:
                continue
            pct = (v - rv[i-1]) / abs(rv[i-1])
            cell = ws.cell(cur, C_QTR0 + i, round(pct, 4))
            cell.font = F(size=8, color="276221" if pct >= 0 else "9C0006")
            cell.fill = BG("D6EEF8"); cell.alignment = AL(); cell.border = BD()
            cell.number_format = '0.0%'
        arv = list(ann_rev.values)
        for i, v in enumerate(arv):
            if i == 0 or not pd.notna(v) or not pd.notna(arv[i-1]) or arv[i-1] == 0:
                continue
            pct = (v - arv[i-1]) / abs(arv[i-1])
            cell = ws.cell(cur, C_YR0 + i, round(pct, 4))
            cell.font = F(True, 8, "276221" if pct >= 0 else "9C0006")
            cell.fill = BG("C4E0F0"); cell.alignment = AL(); cell.border = BD()
            cell.number_format = '0.0%'

        ws.row_dimensions[cur].height = 16; cur += 1

        # ── EPS QoQ % row ─────────────────────────────────────────
        c = ws.cell(cur, C_LBL, "EPS QoQ %")
        c.font = F(True, 8, "276221"); c.fill = BG("D5F5E3")
        c.alignment = AL("left"); c.border = BD()
        ws.cell(cur, C_SEP).fill = BG("F0F4FA")

        ev = list(qtr_eps.values)
        for i, v in enumerate(ev):
            if i == 0 or not pd.notna(v) or not pd.notna(ev[i-1]) or ev[i-1] == 0:
                continue
            pct = (v - ev[i-1]) / abs(ev[i-1])
            cell = ws.cell(cur, C_QTR0 + i, round(pct, 4))
            cell.font = F(size=8, color="276221" if pct >= 0 else "9C0006")
            cell.fill = BG("D5F5E3"); cell.alignment = AL(); cell.border = BD()
            cell.number_format = '0.0%'
        aev = list(ann_eps.values)
        for i, v in enumerate(aev):
            if i == 0 or not pd.notna(v) or not pd.notna(aev[i-1]) or aev[i-1] == 0:
                continue
            pct = (v - aev[i-1]) / abs(aev[i-1])
            cell = ws.cell(cur, C_YR0 + i, round(pct, 4))
            cell.font = F(True, 8, "276221" if pct >= 0 else "9C0006")
            cell.fill = BG("A9DFBF"); cell.alignment = AL(); cell.border = BD()
            cell.number_format = '0.0%'

        ws.row_dimensions[cur].height = 16; cur += 1

        # ── Spacer (chart breathing room) ────────────────────────
        for _ in range(CHART_SPACER):
            ws.row_dimensions[cur].height = 8; cur += 1

        def _bar(title, data_row, cat_col0, cat_col1, anchor, y_fmt='#,##0.00'):
            ch = BarChart()
            ch.type = "col"; ch.grouping = "clustered"; ch.style = 10
            ch.title = title; ch.legend = None
            ch.y_axis.numFmt = y_fmt
            ch.add_data(Reference(ws, min_col=cat_col0, max_col=cat_col1,
                                  min_row=data_row, max_row=data_row), from_rows=True)
            ch.set_categories(Reference(ws, min_col=cat_col0, max_col=cat_col1,
                                        min_row=col_hdr_row))
            ch.width = CHART_W; ch.height = CHART_H
            ch.anchor = anchor
            ws.add_chart(ch)

        q_rev = len(qtr_rev); a_rev = len(ann_rev)
        q_eps = len(qtr_eps); a_eps = len(ann_eps)

        if q_rev > 0:
            _bar(f"{ticker}  |  Doanh thu quý ($B)",
                 rev_row, C_QTR0, C_QTR0+q_rev-1, f"{CL(C_CH1)}{stock_row}")
        if a_rev > 0:
            _bar(f"{ticker}  |  Doanh thu năm ($B)",
                 rev_row, C_YR0, C_YR0+a_rev-1, f"{CL(C_CH2)}{stock_row}")
        if q_eps > 0:
            _bar(f"{ticker}  |  EPS quý ($)",
                 eps_row, C_QTR0, C_QTR0+q_eps-1, f"{CL(C_CH3)}{stock_row}")
        if a_eps > 0:
            _bar(f"{ticker}  |  EPS năm ($)",
                 eps_row, C_YR0, C_YR0+a_eps-1, f"{CL(C_CH4)}{stock_row}")
        if qtr_roe:
            _bar(f"{ticker}  |  ROE quý (ann.)",
                 roe_row, C_QTR0, C_QTR0+N_QTR-1, f"{CL(C_CH5)}{stock_row}", y_fmt='0%')
        if ann_roe:
            _bar(f"{ticker}  |  ROE năm",
                 roe_row, C_YR0, C_YR0+N_YR-1, f"{CL(C_CH6)}{stock_row}", y_fmt='0%')
        if qtr_cfps:
            _bar(f"{ticker}  |  CF/Share quý ($)",
                 cfps_row, C_QTR0, C_QTR0+N_QTR-1, f"{CL(C_CH7)}{stock_row}")
        if ann_cfps:
            _bar(f"{ticker}  |  CF/Share năm ($)",
                 cfps_row, C_YR0, C_YR0+N_YR-1, f"{CL(C_CH8)}{stock_row}")

    print(f"  📊 Financials done ({total} stocks)          ")


def _dashboard_sheet(wb, df):
    from openpyxl.formatting.rule import DataBarRule

    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_tab_color = "1F3864"

    HAS_QC   = "QC_Score" in df.columns and "QC_Signal" in df.columns
    N_QC_LOC = 6

    for ci in range(1, 25):
        ws.column_dimensions[CL(ci)].width = 9
    ws.column_dimensions["C"].width = 20

    def _sec(row, text, bg="2C4F7C"):
        ws.merge_cells(f"A{row}:X{row}")
        c = ws.cell(row, 1, text)
        c.font = F(True, 10, "FFFFFF"); c.fill = BG(bg); c.alignment = AL()
        ws.row_dimensions[row].height = 18

    def _kpi(r, sc, val, lbl, bg, fg="FFFFFF"):
        ec = sc + 3
        ws.merge_cells(f"{CL(sc)}{r}:{CL(ec)}{r}")
        c = ws.cell(r, sc, val)
        c.font = F(True, 18, fg); c.fill = BG(bg); c.alignment = AL()
        ws.row_dimensions[r].height = 36
        ws.merge_cells(f"{CL(sc)}{r+1}:{CL(ec)}{r+1}")
        c2 = ws.cell(r+1, sc, lbl)
        c2.font = F(False, 8, fg); c2.fill = BG(bg); c2.alignment = AL()
        ws.row_dimensions[r+1].height = 16

    # Row 1: Title + Timestamp
    ws.merge_cells("A1:R1")
    c = ws["A1"]
    c.value = "STOCK SCREENER  ·  DECISION DASHBOARD"
    c.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    c.fill = BG("0D2137"); c.alignment = AL()
    ws.row_dimensions[1].height = 36
    ws.merge_cells("S1:X1")
    c2 = ws.cell(1, 19)
    c2.value = f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c2.font = F(False, 8, "88AACC", italic=True)
    c2.fill = BG("0D2137"); c2.alignment = AL("right")
    ws.freeze_panes = ws["A2"]

    # ❶ CAN SLIM KPIs (rows 2-4)
    _sec(2, "❶  CAN SLIM  —  KEY METRICS")
    total   = len(df)
    sb_cnt  = int((df["CS_Signal"] == "\U0001f7e2 STRONG BUY").sum())
    b_cnt   = int((df["CS_Signal"] == "\U0001f535 BUY").sum())
    avg_cs  = f"{df['CS_Score'].mean():.1f}" if "CS_Score" in df.columns else "—"
    _1y     = df["1Y%"].dropna()
    avg_1y  = f"{_1y.mean():.1f}%" if len(_1y) else "—"
    _roe    = df["ROE%"].dropna()
    avg_roe = f"{_roe.mean():.1f}%" if len(_roe) else "—"
    _kpi(3, 1,  str(total),                    "Total Stocks",            "1B3A5C")
    _kpi(3, 5,  f"{sb_cnt/total*100:.1f}%",    f"Strong Buy  ({sb_cnt})", "00703A")
    _kpi(3, 9,  f"{b_cnt/total*100:.1f}%",     f"Buy  ({b_cnt})",         "0070C0")
    _kpi(3, 13, avg_cs,                         f"Avg CS Score / {N_CS}",  "5B2C6F")
    _kpi(3, 17, avg_1y,                         "Avg 1Y Return",           "7D3C0A")
    _kpi(3, 21, avg_roe,                        "Avg ROE%",                "0E6655")
    ws.row_dimensions[5].height = 6

    # ❷ QC KPIs (rows 6-8)
    _sec(6, "❷  QUALITY COMPOUNDER  —  KEY METRICS")
    if HAS_QC:
        comp_cnt  = int((df["QC_Signal"] == "\U0001f3c6 COMPOUNDER").sum())
        qual_cnt  = int((df["QC_Signal"] == "⭐ QUALITY").sum())
        avg_qc_v  = df["QC_Score"].mean()
        avg_qc_s  = f"{avg_qc_v:.1f}" if pd.notna(avg_qc_v) else "—"
        _roic     = df["ROIC%"].dropna() if "ROIC%" in df.columns else pd.Series(dtype=float)
        avg_roic  = f"{_roic.mean():.1f}%" if len(_roic) else "—"
        dual_cnt  = int(((df["CS_Score"] >= 7) & (df["QC_Score"] >= 4)).sum())
        _qc_mask  = df["QC_Signal"].isin(["\U0001f3c6 COMPOUNDER", "⭐ QUALITY"])
        eq_cb_cnt = (int((df[_qc_mask]["EQ_Badge"] == "\U0001f49a Cash Backed").sum())
                     if "EQ_Badge" in df.columns else 0)
    else:
        comp_cnt = qual_cnt = dual_cnt = eq_cb_cnt = 0
        avg_qc_s = avg_roic = "—"
    _kpi(7, 1,  str(comp_cnt), "# COMPOUNDER",                "1A5C2B")
    _kpi(7, 5,  str(qual_cnt), "# QUALITY",                   "1A3A5C")
    _kpi(7, 9,  avg_roic,      "Avg ROIC%",                   "0E6655")
    _kpi(7, 13, avg_qc_s,      f"Avg QC Score / {N_QC_LOC}", "5B2C6F")
    _kpi(7, 17, str(dual_cnt), "Dual Leaders (CS≥7 & QC≥4)", "4A235A")
    _kpi(7, 21, str(eq_cb_cnt), "\U0001f49a Cash Backed (QC stocks)", "155A28")
    ws.row_dimensions[9].height = 6

    # RISK FLAGS (row 10)
    risk_de = int((df["D/E"].fillna(0)  > 1).sum())  if "D/E"  in df.columns else 0
    risk_pe = int((df["P/E"].fillna(0)  > 50).sum()) if "P/E"  in df.columns else 0
    risk_1m = int((df["1M%"].fillna(0) < -10).sum()) if "1M%"  in df.columns else 0
    ws.merge_cells("A10:X10")
    c = ws.cell(10, 1,
        f"⚠️  RISK FLAGS:   D/E > 1: {risk_de} mã   |   P/E > 50: {risk_pe} mã   |   1M% < −10%: {risk_1m} mã")
    c.font = F(True, 9, "9C6500"); c.fill = BG("FFF2CC"); c.alignment = AL()
    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 8

    # 📊 TOP CS SCORE + TOP QC SCORE — side by side (row 12+)
    # CS: cols 1-7 | spacer: col 8 | QC: cols 9-15
    ws.merge_cells("A12:G12")
    c = ws.cell(12, 1, "📊  TOP CAN SLIM SCORE")
    c.font = F(True, 10, "FFFFFF"); c.fill = BG("1B3A5C"); c.alignment = AL()
    ws.cell(12, 8).fill = BG("F0F4FA")
    ws.merge_cells("I12:O12")
    c = ws.cell(12, 9, "📊  TOP QUALITY COMPOUNDER SCORE")
    c.font = F(True, 10, "FFFFFF"); c.fill = BG("0E6655"); c.alignment = AL()
    ws.row_dimensions[12].height = 18

    CS_HDRS = ["#", "Ticker", "Company", "Sector", "CS Score", "Signal", "1Y%"]
    CS_WDS  = [4, 8, 20, 14, 8, 14, 8]
    for ci, (h, w) in enumerate(zip(CS_HDRS, CS_WDS), 1):
        ws.column_dimensions[CL(ci)].width = w
        c = ws.cell(13, ci, h)
        c.font = F(True, 9, "FFFFFF"); c.fill = BG("1B3A5C")
        c.alignment = AL(); c.border = BD()

    QC_HDRS = ["#", "Ticker", "Company", "Sector", "QC Score", "Signal", "ROIC%"]
    QC_WDS  = [4, 8, 20, 14, 8, 14, 8]
    QC_C0   = 9
    for ci, (h, w) in enumerate(zip(QC_HDRS, QC_WDS), QC_C0):
        ws.column_dimensions[CL(ci)].width = w
        c = ws.cell(13, ci, h)
        c.font = F(True, 9, "FFFFFF"); c.fill = BG("0E6655")
        c.alignment = AL(); c.border = BD()
    ws.column_dimensions["H"].width = 2
    ws.row_dimensions[13].height = 18

    _cs_buy = ["\U0001f7e2 STRONG BUY", "\U0001f535 BUY"]
    cs_top = (df[df["CS_Signal"].isin(_cs_buy)]
              .sort_values("CS_Score", ascending=False)
              if "CS_Score" in df.columns else pd.DataFrame())
    _qc_buy = ["\U0001f3c6 COMPOUNDER", "⭐ QUALITY"]
    qc_top = (df[df["QC_Signal"].isin(_qc_buy)]
              .sort_values("QC_Score", ascending=False)
              if HAS_QC else pd.DataFrame())
    n_rows = max(len(cs_top), len(qc_top), 1)

    cs_end = 13
    for ri in range(1, n_rows + 1):
        er  = 13 + ri
        ws.row_dimensions[er].height = 17

        # CS side
        if ri <= len(cs_top):
            row  = cs_top.iloc[ri - 1]
            rbg  = "EBF5FB" if ri % 2 == 0 else "F4F9FD"
            sig  = row.get("CS_Signal", "") or ""
            fg_s, bg_s = SIGNAL_STYLE.get(sig, ("000000", "FFFFFF"))
            def _cw(ci, val, bold=False, fg="000000", bgc=None, align="center", nfmt=None):
                c = ws.cell(er, ci, val)
                c.border = BD(); c.alignment = AL(align)
                c.fill = BG(bgc if bgc else rbg); c.font = F(bold, 9, fg)
                if nfmt and val is not None: c.number_format = nfmt
            _cw(1, ri,  True, "FFFFFF", "1B3A5C")
            _cw(2, row.get("Ticker", ""),             True,  "1B3A5C")
            _cw(3, row.get("T\xean C\xf4ng Ty", ""), False, "1C3550", align="left")
            _cw(4, row.get("Sector", ""),             False, "334466", align="left")
            _cw(5, int(row.get("CS_Score", 0) or 0), True,  CG_FG, CG_BG)
            _cw(6, sig, True, fg_s, bg_s)
            v1y = row.get("1Y%")
            if v1y is not None and not (isinstance(v1y, float) and pd.isna(v1y)):
                c7 = ws.cell(er, 7); c7.value = float(v1y) / 100
                c7.number_format = '+0.0%;(0.0%);"-"'; c7.border = BD(); c7.fill = BG(rbg)
                c7.alignment = AL(); c7.font = F(size=9, color="276221" if float(v1y) > 0 else "9C0006")
            else:
                _cw(7, "—", fg="BBBBBB")

        # spacer col 8
        ws.cell(er, 8).fill = BG("F0F4FA")

        # QC side
        if ri <= len(qc_top):
            row  = qc_top.iloc[ri - 1]
            rbgq = "E8F8F5" if ri % 2 == 0 else "F0FBF9"
            sig  = row.get("QC_Signal", "") or ""
            fg_q, bg_q = QC_SIGNAL_STYLE.get(sig, ("000000", "FFFFFF"))
            def _qw(ci, val, bold=False, fg="000000", bgc=None, align="center", nfmt=None):
                c = ws.cell(er, ci, val)
                c.border = BD(); c.alignment = AL(align)
                c.fill = BG(bgc if bgc else rbgq); c.font = F(bold, 9, fg)
                if nfmt and val is not None: c.number_format = nfmt
            _qw(QC_C0,   ri,  True, "FFFFFF", "0E6655")
            _qw(QC_C0+1, row.get("Ticker", ""),             True,  "0E6655")
            _qw(QC_C0+2, row.get("T\xean C\xf4ng Ty", ""), False, "1C3550", align="left")
            _qw(QC_C0+3, row.get("Sector", ""),             False, "334466", align="left")
            _qw(QC_C0+4, int(row.get("QC_Score", 0) or 0), True,  CB_FG, CB_BG)
            _qw(QC_C0+5, sig, True, fg_q, bg_q)
            roic = row.get("ROIC%")
            if roic is not None and not (isinstance(roic, float) and pd.isna(roic)):
                cr = ws.cell(er, QC_C0+6); cr.value = float(roic) / 100
                cr.number_format = '0.0%'; cr.border = BD(); cr.fill = BG(rbgq)
                cr.alignment = AL()
                cr.font = F(True, 9, "276221" if float(roic) >= 15 else "0070C0")
            else:
                _qw(QC_C0+6, "—", fg="BBBBBB")
        elif not HAS_QC and ri == 1:
            ws.merge_cells(f"I{er}:O{er}")
            c = ws.cell(er, QC_C0, "— Y\xeau cầu bật yfinance —")
            c.font = F(False, 8, "888888", italic=True); c.alignment = AL()

        cs_end = er
    ws.row_dimensions[cs_end + 1].height = 8

    # ⭐ DUAL LEADERS (dynamic, after cs_end)
    dl_sec = cs_end + 2
    _sec(dl_sec, "⭐  DUAL LEADERS  —  CS Score ≥ 7  AND  QC Score ≥ 4", bg="4A235A")
    DL_HDRS = ["#", "Ticker", "Company", "Sector", "Price", "CS", "QC",
               "1Y%", "ROE%", "ROIC%", "D/E", "Signal", "⚠"]
    ws.column_dimensions["M"].width = 11
    dl_hdr_row = dl_sec + 1
    for ci, h in enumerate(DL_HDRS, 1):
        c = ws.cell(dl_hdr_row, ci, h)
        c.font = F(True, 9, "FFFFFF"); c.fill = BG("4A235A")
        c.alignment = AL(); c.border = BD()
    ws.row_dimensions[dl_hdr_row].height = 18

    dl_df = (df[(df["CS_Score"] >= 7) & (df["QC_Score"] >= 4)]
             .sort_values(["CS_Score", "QC_Score"], ascending=[False, False])
             if HAS_QC else pd.DataFrame())

    dl_end = dl_hdr_row
    if dl_df.empty:
        ws.merge_cells(f"A{dl_hdr_row+1}:M{dl_hdr_row+1}")
        c = ws.cell(dl_hdr_row + 1, 1, "— Kh\xf4ng c\xf3 m\xe3 n\xe0o đạt cả 2 ti\xeau ch\xed —")
        c.font = F(False, 9, "888888", italic=True); c.alignment = AL()
        ws.row_dimensions[dl_hdr_row + 1].height = 18; dl_end = dl_hdr_row + 1
    else:
        for ri, (_, row) in enumerate(dl_df.iterrows(), 1):
            er  = dl_hdr_row + ri
            rbg = "F5F0FF" if ri % 2 == 0 else "FBF7FF"
            ws.row_dimensions[er].height = 17
            sig      = row.get("CS_Signal", "") or ""
            fg_s, bg_s = SIGNAL_STYLE.get(sig, ("000000", "FFFFFF"))

            def _w(ci, val, bold=False, fg="000000", bgc=None, align="center", nfmt=None):
                c = ws.cell(er, ci, val)
                c.border = BD(); c.alignment = AL(align)
                c.fill = BG(bgc if bgc else rbg); c.font = F(bold, 9, fg)
                if nfmt and val is not None: c.number_format = nfmt

            _w(1, ri,   True,  "FFFFFF", "4A235A")
            _w(2, row.get("Ticker", ""),      True,  "1B3A5C")
            _w(3, row.get("T\xean C\xf4ng Ty", ""), False, "1C3550", align="left")
            _w(4, row.get("Sector", ""),      False, "334466", align="left")
            p = row.get("Price ($)")
            if p is not None and not (isinstance(p, float) and pd.isna(p)):
                _w(5, p, nfmt='"$"#,##0.00')
            else:
                _w(5, "—", fg="BBBBBB")
            _w(6, int(row.get("CS_Score", 0) or 0), True, CG_FG, CG_BG)
            _w(7, int(row.get("QC_Score", 0) or 0), True, CB_FG, CB_BG)
            for ci_n, key, mul, nfmt in [
                (8,  "1Y%",   100, '+0.0%;(0.0%);"-"'),
                (9,  "ROE%",  100, '0.0%'),
                (10, "ROIC%", 100, '0.0%'),
                (11, "D/E",     1, '0.0'),
            ]:
                v = row.get(key)
                c = ws.cell(er, ci_n)
                c.border = BD(); c.alignment = AL(); c.fill = BG(rbg)
                if v is not None and not (isinstance(v, float) and pd.isna(v)):
                    fv = float(v); c.value = fv / mul; c.number_format = nfmt
                    if   ci_n == 8:  c.font = F(size=9, color="276221" if fv > 0 else "9C0006")
                    elif ci_n == 9:  c.font = F(size=9, color="276221" if fv > 17 else "000000")
                    elif ci_n == 10: c.font = F(True, 9, "276221" if fv >= 15 else "0070C0" if fv >= 10 else "7D6608")
                    else:            c.font = F(size=9, color="9C0006" if fv > 2 else "7D6608" if fv > 1 else "276221")
                else:
                    c.value = "—"; c.font = F(size=9, color="BBBBBB")
            _w(12, sig, True, fg_s, bg_s)
            de_v = row.get("D/E"); pe_v = row.get("P/E")
            flags = []
            if de_v is not None and not (isinstance(de_v, float) and pd.isna(de_v)) and float(de_v) > 1:
                flags.append("D/E")
            if pe_v is not None and not (isinstance(pe_v, float) and pd.isna(pe_v)) and float(pe_v) > 50:
                flags.append("P/E")
            if flags:
                _w(13, "⚠️ " + " · ".join(flags), True, "9C6500", "FFF2CC")
            else:
                _w(13, "", bgc=rbg)
            dl_end = er
        ws.conditional_formatting.add(
            f"H{dl_hdr_row+1}:H{dl_end}",
            DataBarRule(start_type="min", start_value=0, end_type="max", end_value=100, color="0070C0"))

    # 🎯 TOP PICKS — 4 mini-tables (side by side, dynamic height)
    ws.row_dimensions[dl_end + 1].height = 10
    tp_sec = dl_end + 2
    _sec(tp_sec, "\U0001f3af  TOP PICKS", bg="1A3A5C")
    tp_hdr = tp_sec + 1
    # col offsets: Momentum=1, Quality=7, Value=13, Breakout=19
    TP_GRP = [
        (1,  "\U0001f680 Momentum  (STRONG BUY, 1Y% ↓)",     "1A5276"),
        (7,  "\U0001f48e Quality  (Compounder, ROIC% ↓)",     "1A5C2B"),
        (13, "\U0001f4c9 Value  (P/E<35, EPS>10%, CS≥6)",     "784212"),
        (19, "\U0001f4a5 Breakout  (52W High ≥90%, CS≥7)",    "4A235A"),
    ]
    COL_HDRS = ["Ticker", "Metric", "Sector"]
    TP_BG    = ["1A5276", "1A5C2B", "784212", "4A235A"]

    def _tp_rows():
        # Momentum: STRONG BUY sort 1Y%; fallback thêm BUY nếu trống
        mom = []
        if "1Y%" in df.columns and "CS_Signal" in df.columns:
            for sig_filter in ["\U0001f7e2 STRONG BUY",
                               ["\U0001f7e2 STRONG BUY", "\U0001f535 BUY"]]:
                mask = (df["CS_Signal"] == sig_filter if isinstance(sig_filter, str)
                        else df["CS_Signal"].isin(sig_filter))
                d = (df[mask][["Ticker","1Y%","Sector","CS_Signal"]]
                     .dropna(subset=["1Y%"]).sort_values("1Y%", ascending=False))
                if not d.empty:
                    mom = [(r["Ticker"],
                            f"{r['1Y%']:+.1f}%  ({r['CS_Signal'].split()[1]})",
                            r.get("Sector",""))
                           for _, r in d.iterrows()]
                    break

        # Quality: Compounder sort ROIC%
        qua = []
        if HAS_QC and "ROIC%" in df.columns:
            d = (df[df["QC_Signal"] == "\U0001f3c6 COMPOUNDER"][["Ticker","ROIC%","Sector"]]
                 .dropna(subset=["ROIC%"]).sort_values("ROIC%", ascending=False))
            qua = [(r["Ticker"], f"ROIC {r['ROIC%']:.1f}%", r.get("Sector",""))
                   for _, r in d.iterrows()]

        # Value: P/E<35 + EPS growth + CS≥6, sort P/E asc
        val = []
        if "P/E" in df.columns and "EPS Annual%" in df.columns and "CS_Score" in df.columns:
            tmp = (df[
                (df["CS_Score"] >= 6) &
                (df["P/E"] > 0) & (df["P/E"] < 35) &
                (df["EPS Annual%"] > 10) & (df["EPS Annual%"] < 200)
            ].dropna(subset=["P/E"]).sort_values("P/E"))
            val = [(r["Ticker"],
                    f"P/E {r['P/E']:.1f}× | EPS +{r['EPS Annual%']:.0f}%",
                    r.get("Sector",""))
                   for _, r in tmp.iterrows()]

        # Breakout: gần 52W High (≥90%) + CS Score cao
        brk = []
        if "52W_High%" in df.columns and "CS_Score" in df.columns:
            if df["52W_High%"].isna().all():
                brk = [("—", "52W High% N/A — chạy không có --no-yf để lấy dữ liệu", "yfinance off")]
            else:
                tmp = (df[
                    (df["52W_High%"] >= 90) &
                    (df["CS_Score"] >= 7)
                ].sort_values(["CS_Score", "52W_High%"], ascending=[False, False]))
                brk = [(r["Ticker"],
                        f"52W {r['52W_High%']:.1f}% | CS {int(r['CS_Score'])}",
                        r.get("Sector",""))
                       for _, r in tmp.iterrows()]

        return mom, qua, val, brk

    tp_data = _tp_rows()
    max_tp_rows = max((len(rows) for rows in tp_data), default=1)

    for (sc, title, _), bg_h in zip(TP_GRP, TP_BG):
        ws.merge_cells(f"{CL(sc)}{tp_hdr}:{CL(sc+2)}{tp_hdr}")
        c = ws.cell(tp_hdr, sc, title)
        c.font = F(True, 9, "FFFFFF"); c.fill = BG(bg_h); c.alignment = AL()
        ws.row_dimensions[tp_hdr].height = 18
        for i, h in enumerate(COL_HDRS):
            c2 = ws.cell(tp_hdr + 1, sc + i, h)
            c2.font = F(True, 8, "FFFFFF"); c2.fill = BG(bg_h)
            c2.alignment = AL(); c2.border = BD()
        ws.row_dimensions[tp_hdr + 1].height = 16

    for (sc, _, _), bg_h, rows in zip(TP_GRP, TP_BG, tp_data):
        if not rows:
            ws.merge_cells(f"{CL(sc)}{tp_hdr+2}:{CL(sc+2)}{tp_hdr+2}")
            c = ws.cell(tp_hdr + 2, sc, "— No data —")
            c.font = F(False, 8, "888888", italic=True); c.alignment = AL()
            continue
        for ri, (ticker, metric, sector) in enumerate(rows):
            er2 = tp_hdr + 2 + ri
            rbg2 = "F0F8FF" if ri % 2 == 0 else "FFFFFF"
            ws.row_dimensions[er2].height = 16
            for ci_off, val, bold, fg in [
                (0, ticker, True,  "1B3A5C"),
                (1, metric, True,  "000000"),
                (2, sector, False, "334466"),
            ]:
                c = ws.cell(er2, sc + ci_off, val)
                c.fill = BG(rbg2); c.border = BD()
                c.font = F(bold, 9, fg)
                c.alignment = AL("left" if ci_off == 2 else "center")

    tp_end = tp_hdr + 1 + max_tp_rows + 1   # dynamic, không hardcode

    # Pre-compute zone data so ch_end can be set dynamically
    zone_data = []; watch_n = 0
    if HAS_QC:
        _cs = df["CS_Score"]; _qc = df["QC_Score"]
        _dual = df[(_cs >= 7) & (_qc >= 4)].sort_values(["CS_Score","QC_Score"], ascending=[False,False])
        _mom  = df[(_cs >= 7) & (_qc <  4)].sort_values(["CS_Score","1Y%"],      ascending=[False,False])
        _qual = df[(_cs <  7) & (_qc >= 4)].sort_values("QC_Score",              ascending=False)
        watch_n = int(((_cs < 7) & (_qc < 4)).sum())
        zone_data = [
            ("⭐  Dual Leaders  (CS ≥ 7  &  QC ≥ 4)", "4A235A", _dual),
            ("🚀  Momentum  (CS ≥ 7,  QC < 4)",        "1A5276", _mom),
            ("💎  Quality  (CS < 7,  QC ≥ 4)",         "1A5C2B", _qual),
        ]
    if zone_data:
        _top_n = max(len(zone_data[0][2]), len(zone_data[2][2]))  # Dual vs Quality
        _bot_n = max(1, len(zone_data[1][2]))                     # Momentum
        zone_rows = 4 + _top_n + _bot_n  # top(hdr+col+data) + divider + bottom(hdr+col+data)
    else:
        zone_rows = 10
    zone_rows = max(zone_rows, 22)

    # ❸ CHART ANALYSIS
    ws.row_dimensions[tp_end].height = 10
    ch_sec = tp_end + 1
    _sec(ch_sec, "❸  CHART ANALYSIS")
    ch_start = ch_sec + 1
    for r in range(ch_start, ch_start + zone_rows + 2):
        ws.row_dimensions[r].height = 17
    ch_end = ch_start + zone_rows

    # ❹ SECTOR BREAKDOWN
    ws.row_dimensions[ch_end + 1].height = 10
    s4_row = ch_end + 2
    _sec(s4_row, "❹  SECTOR BREAKDOWN  —  Sort: # Strong Buy ↓")
    S4_HDRS = ["Sector", "# Stocks", "# Str.Buy", "# Buy", "# Comp.", "Avg CS", "Avg QC", "Avg 1Y%"]
    S4_WD   = [16, 8, 9, 7, 8, 7, 7, 8]
    for ci, (h, w) in enumerate(zip(S4_HDRS, S4_WD), 1):
        ws.column_dimensions[CL(ci)].width = w
        c = ws.cell(s4_row + 1, ci, h)
        c.font = F(True, 9, "FFFFFF"); c.fill = BG("2C4F7C")
        c.alignment = AL(); c.border = BD()
    ws.row_dimensions[s4_row + 1].height = 18

    sec_pivot = []
    for sec, g in df.assign(Sector=df["Sector"].fillna("Unknown")).groupby("Sector"):
        sec_pivot.append({
            "s":   sec,
            "n":   len(g),
            "sb":  int((g["CS_Signal"] == "\U0001f7e2 STRONG BUY").sum()),
            "b":   int((g["CS_Signal"] == "\U0001f535 BUY").sum()),
            "cp":  int((g["QC_Signal"] == "\U0001f3c6 COMPOUNDER").sum()) if HAS_QC else 0,
            "acs": g["CS_Score"].mean() if "CS_Score" in g.columns else None,
            "aqc": g["QC_Score"].mean() if HAS_QC else None,
            "a1y": g["1Y%"].dropna().mean() if "1Y%" in g.columns else None,
        })
    sec_pivot.sort(key=lambda x: x["sb"], reverse=True)
    TOP3_BG = ["FFF2CC", "E8F5E9", "E3F2FD"]
    for ri, row in enumerate(sec_pivot):
        er = s4_row + 2 + ri
        rbg = TOP3_BG[ri] if ri < 3 else ("F0F4FA" if ri % 2 == 0 else "FFFFFF")
        ws.row_dimensions[er].height = 17
        c = ws.cell(er, 1, row["s"]); c.border=BD(); c.fill=BG(rbg); c.alignment=AL("left"); c.font=F(False,9,"1B3A5C")
        for ci, key in [(2,"n"),(3,"sb"),(4,"b"),(5,"cp")]:
            v = row[key]; is_sb = ci == 3 and v > 0
            c = ws.cell(er, ci, v); c.border=BD(); c.fill=BG(rbg); c.alignment=AL()
            c.font = F(True if is_sb else False, 9, "276221" if is_sb else "000000")
        for ci, key in [(6,"acs"),(7,"aqc")]:
            v = row[key]
            c = ws.cell(er, ci); c.border=BD(); c.fill=BG(rbg); c.alignment=AL()
            c.value = round(v, 1) if v is not None and not pd.isna(v) else "—"
            c.font = F(False, 9, "000000" if v is not None and not pd.isna(v) else "BBBBBB")
        v8 = row["a1y"]
        c = ws.cell(er, 8); c.border=BD(); c.fill=BG(rbg); c.alignment=AL()
        if v8 is not None and not pd.isna(v8):
            c.value = v8 / 100; c.number_format = '+0.0%;(0.0%);"-"'
            c.font = F(False, 9, "276221" if v8 > 0 else "9C0006")
        else:
            c.value = "—"; c.font = F(size=9, color="BBBBBB")

    # ZONE 2×2 QUADRANT TABLE — mirrors CS×QC scatter layout
    # Layout:  Left cols (LC) = CS<7  |  divider  |  Right cols (RC) = CS≥7
    #          Top rows = QC≥4  |  thick divider  |  Bottom rows = QC<4
    LC = [1, 2, 3, 4]   # Ticker | CS | QC | 1Y%
    DC = 5               # divider column (uses sector width, just colored gray)
    RC = [6, 7, 8, 9]   # Ticker | CS | QC | 1Y%
    ws.column_dimensions["I"].width = 8   # RC 1Y%

    def _zh(r, cols, label, bg):
        ws.merge_cells(f"{CL(cols[0])}{r}:{CL(cols[-1])}{r}")
        c = ws.cell(r, cols[0], label)
        c.font = F(True, 10, "FFFFFF"); c.fill = BG(bg)
        c.alignment = AL("left"); c.border = BD()
        ws.row_dimensions[r].height = 18

    def _ch(r, cols):
        for ci, h in zip(cols, ["Ticker", "CS", "QC", "1Y%"]):
            c = ws.cell(r, ci, h)
            c.font = F(True, 9, "FFFFFF"); c.fill = BG("2C4F7C")
            c.alignment = AL(); c.border = BD()
        ws.row_dimensions[r].height = 16

    def _div(r):
        c = ws.cell(r, DC); c.fill = BG("C5D5E8"); c.border = BD()

    def _stock(r, cols, row, rbg):
        c = ws.cell(r, cols[0], str(row.get("Ticker", "")))
        c.border = BD(); c.fill = BG(rbg); c.alignment = AL("left"); c.font = F(True, 9, "1B3A5C")
        c = ws.cell(r, cols[1], int(row.get("CS_Score", 0) or 0))
        c.border = BD(); c.fill = BG(CG_BG); c.alignment = AL(); c.font = F(True, 9, CG_FG)
        qc_v = row.get("QC_Score")
        c = ws.cell(r, cols[2])
        c.border = BD(); c.alignment = AL()
        if qc_v is not None and not (isinstance(qc_v, float) and pd.isna(qc_v)):
            c.value = int(qc_v); c.fill = BG(CB_BG); c.font = F(True, 9, CB_FG)
        else:
            c.value = "—"; c.fill = BG(rbg); c.font = F(size=9, color="BBBBBB")
        v = row.get("1Y%")
        c = ws.cell(r, cols[3])
        c.border = BD(); c.alignment = AL(); c.fill = BG(rbg)
        if v is not None and not (isinstance(v, float) and pd.isna(v)):
            c.value = float(v) / 100; c.number_format = '+0.0%;(0.0%)'
            c.font = F(size=9, color="276221" if float(v) > 0 else "9C0006")
        else:
            c.value = "—"; c.font = F(size=9, color="BBBBBB")
        ws.row_dimensions[r].height = 17

    def _pad(r, cols, bg):
        for ci in cols:
            c = ws.cell(r, ci); c.border = BD(); c.fill = BG(bg)

    if not zone_data:
        ws.merge_cells(f"A{ch_start}:I{ch_start}")
        c = ws.cell(ch_start, 1, "— Bật yfinance Moat để xem phân vùng CS×QC —")
        c.font = F(False, 10, "888888", italic=True); c.fill = BG("F5F5F5"); c.alignment = AL()
    else:
        _dual_df = zone_data[0][2]; _mom_df = zone_data[1][2]; _qual_df = zone_data[2][2]
        top_n = max(len(_qual_df), len(_dual_df))
        bot_n = max(1, len(_mom_df))
        qual_rows = list(_qual_df.iterrows()); dual_rows = list(_dual_df.iterrows())
        mom_rows  = list(_mom_df.iterrows())

        # ── TOP HALF: 💎 Quality (left) | ⭐ Dual Leaders (right) ──
        _zh(ch_start,     LC, "💎  Quality  (CS < 7,  QC ≥ 4)",        "1A5C2B")
        _zh(ch_start,     RC, "⭐  Dual Leaders  (CS ≥ 7  &  QC ≥ 4)", "4A235A")
        _div(ch_start)
        _ch(ch_start + 1, LC); _ch(ch_start + 1, RC); _div(ch_start + 1)
        ws.row_dimensions[ch_start + 1].height = 16

        for i in range(top_n):
            r = ch_start + 2 + i; _div(r)
            if i < len(qual_rows):
                _stock(r, LC, qual_rows[i][1], "EAFAF1" if i % 2 == 0 else "FFFFFF")
            else:
                _pad(r, LC, "F4FDF8")
            if i < len(dual_rows):
                _stock(r, RC, dual_rows[i][1], "F5F0FF" if i % 2 == 0 else "FFFFFF")
            else:
                _pad(r, RC, "FAF7FF")

        # ── HORIZONTAL DIVIDER ──
        div_r = ch_start + 2 + top_n
        for ci in LC + [DC] + RC:
            c = ws.cell(div_r, ci); c.fill = BG("334466"); c.border = BD()
        ws.row_dimensions[div_r].height = 5

        # ── BOTTOM HALF: 👀 Watchlist (left) | 🚀 Momentum (right) ──
        b0 = div_r + 1
        _zh(b0, LC, "👀  Watchlist  (CS < 7  &  QC < 4)", "555555")
        _zh(b0, RC, "🚀  Momentum  (CS ≥ 7,  QC < 4)",    "1A5276")
        _div(b0)

        # Watchlist — merged cell spanning col-header + all data rows
        ws.merge_cells(f"{CL(LC[0])}{b0+1}:{CL(LC[-1])}{b0+1+bot_n}")
        c = ws.cell(b0 + 1, LC[0], f"{watch_n} mã\nXem sheet chính để lọc thêm")
        c.font = F(False, 11, "999999", italic=True); c.fill = BG("F5F5F5")
        c.alignment = AL("center"); c.border = BD()

        # Momentum — col headers at b0+1, data from b0+2
        _ch(b0 + 1, RC); _div(b0 + 1)
        for i, (_, row) in enumerate(mom_rows):
            r = b0 + 2 + i; _div(r)
            _stock(r, RC, row, "EBF5FB" if i % 2 == 0 else "FFFFFF")
        for i in range(len(mom_rows), bot_n):
            r = b0 + 2 + i; _div(r); _pad(r, RC, "F0F8FF")


def _main_sheet(wb, df, market, top):
    ws = wb.active; ws.title = f"Top{top}"
    ws.sheet_view.showGridLines = False
    bdr = BD(); nd = len(DATA_HEADERS)
    CS_START    = nd + 1
    CI_SCORE    = CS_START + N_CS
    CI_CONV     = CS_START + N_CS + 1
    CI_SIGNAL   = CS_START + N_CS + 2
    TOTAL       = CI_SIGNAL
    DR = 5

    # Sort theo CS_Score → Conviction giảm dần để mã tốt nhất lên đầu
    sort_cols = [c for c in ["CS_Score", "Conviction"] if c in df.columns]
    df = df.sort_values(sort_cols, ascending=[False] * len(sort_cols)).reset_index(drop=True) if sort_cols else df

    ws.merge_cells(f"A1:{CL(TOTAL)}1")
    c = ws["A1"]; c.value = f"📊  TOP {top} {market.upper()}  —  FUNDAMENTAL + CAN SLIM SCREENER"
    c.font = F(True,15,C_WHITE); c.fill = BG(C_DARK); c.alignment = AL(); ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{CL(TOTAL)}2")
    c = ws["A2"]; c.value = f"Source: TradingView + yfinance (Moat)   |   {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   Sorted by CS Score ↓"
    c.font = F(False,8,"88AACC",True); c.fill = BG(C_DARK); c.alignment = AL(); ws.row_dimensions[2].height = 16

    groups = [
        (1,1,"TICKER","213A5C"),(2,2,"TÊN CÔNG TY","1C3550"),
        (3,3,"SECTOR","2C4F7C"),
        (4,5,"PRICE","2E5E8E"),
        (6,7,"🏰  MOAT (yfinance 5yr)","4A235A"),
        (8,11,"GROWTH","1A5276"),
        (12,15,"QUALITY","154360"),(16,18,"VALUATION","1B2631"),
        (19,23,"PERFORMANCE","0E3251"),
        (CS_START, CI_SCORE-1, "✅  CAN SLIM  (✓ = đạt / ✗ = không đạt)", "1A5C2B"),
        (CI_SCORE, CI_SCORE, "SCORE", "0D3B1A"),
        (CI_CONV, CI_CONV, "CONVICTION", "0D3B1A"),
        (CI_SIGNAL, CI_SIGNAL, "SIGNAL", "0D3B1A"),
    ]
    ws.row_dimensions[3].height = 14
    for cs, ce, lbl, bg in groups:
        c = ws.cell(3, cs, lbl)
        c.font = F(True,8,"C8DDEF"); c.fill = BG(bg); c.alignment = AL(); c.border = bdr
        if cs < ce: ws.merge_cells(f"{CL(cs)}3:{CL(ce)}3")

    ws.row_dimensions[4].height = 30
    for ci, h in enumerate(DATA_HEADERS, 1):
        c = ws.cell(4,ci,h); c.font=F(True,8,C_WHITE); c.fill=BG(C_NAVY); c.alignment=AL(wrap=True); c.border=bdr

    for i, key in enumerate(CS_KEYS):
        ci = CS_START + i
        c = ws.cell(4,ci,CANSLIM[key]["label"])
        c.font=F(True,8,C_WHITE); c.fill=BG("1A5C2B"); c.alignment=AL(wrap=True); c.border=bdr
    for ci, lbl in [(CI_SCORE,f"Score\n/{N_CS}"), (CI_CONV,"Conviction"), (CI_SIGNAL,"Signal")]:
        c = ws.cell(4,ci,lbl); c.font=F(True,9,C_WHITE); c.fill=BG("0D3B1A"); c.alignment=AL(wrap=True); c.border=bdr

    for ri, (_, row) in enumerate(df.iterrows(), 1):
        er = DR + ri - 1; alt = ri%2==0; rbg = C_ALT if alt else C_WHITE

        for ci, h in enumerate(DATA_HEADERS, 1):
            val = row.get(h); c = ws.cell(er,ci)
            c.border=bdr; c.fill=BG(rbg); c.alignment=AL("left" if ci==1 else "center")
            if ci==1:
                c.value=val; c.font=F(True,9,C_NAVY); continue
            if val is None or (isinstance(val,float) and pd.isna(val)):
                c.font=F(size=9,color="BBBBBB"); continue
            if h=="Tên Công Ty":
                c.value=val; c.font=F(size=8,color="1C3550"); c.alignment=AL("left")
            elif h=="Sector":
                c.value=val; c.font=F(size=8,italic=True,color="334466"); c.alignment=AL("left")
            elif h=="Moat Proxy":
                c.value=val; c.font=F(size=8,italic=True,color="4A235A")
                c.fill=BG("F5EEF8"); c.alignment=AL("left")
            elif h=="Moat Score":
                fg,bg = MOAT_SCORE_STYLE.get(val, ("000000",C_WHITE))
                c.value=val; c.font=F(True,9,fg); c.fill=BG(bg); c.alignment=AL()
            elif h=="Price ($)":
                c.value=val; c.number_format='"$"#,##0.00'; c.font=F(size=9)
            elif h=="MCap ($B)":
                c.value=val; c.number_format='#,##0.0'; c.font=F(size=9)
            elif ci in PCT_COLS:
                c.value=val/100; c.number_format='+0.0%;(0.0%);"-"'
                c.font = F(size=9,color=CG_FG,bold=(val>25)) if val>0 else (F(size=9,color=CR_FG) if val<0 else F(size=9))
            else:
                c.value=val; c.number_format='0.0'; c.font=F(size=9)

        for i, key in enumerate(CS_KEYS):
            ci = CS_START + i; v = row.get(f"CS_{key}"); c = ws.cell(er,ci)
            c.border=bdr; c.alignment=AL()
            if v is True:    c.value="✓"; c.font=F(True,11,CG_FG); c.fill=BG(CG_BG)
            elif v is False: c.value="✗"; c.font=F(True,10,CR_FG); c.fill=BG(CR_BG)
            else:            c.value="–"; c.font=F(size=9,color="AAAAAA"); c.fill=BG("F2F2F2")

        sc = row.get("CS_Score",0); c = ws.cell(er,CI_SCORE,sc)
        c.border=bdr; c.alignment=AL(); c.number_format="0"
        pct = sc/N_CS
        if pct>=0.875:  c.fill=BG(CG_BG); c.font=F(True,11,CG_FG)
        elif pct>=0.625: c.fill=BG(CB_BG); c.font=F(True,11,CB_FG)
        elif pct>=0.375: c.fill=BG(CY_BG); c.font=F(True,11,CY_FG)
        else:            c.fill=BG(CR_BG); c.font=F(True,11,CR_FG)

        cv = row.get("Conviction", sc); c = ws.cell(er, CI_CONV, cv)
        c.border=bdr; c.alignment=AL(); c.number_format="0.0"
        pct_cv = cv / (N_CS * 1.2)  # max conviction = N_CS × WIDE moat multiplier
        if pct_cv>=0.75:  c.fill=BG(CG_BG); c.font=F(True,10,CG_FG)
        elif pct_cv>=0.55: c.fill=BG(CB_BG); c.font=F(True,10,CB_FG)
        elif pct_cv>=0.35: c.fill=BG(CY_BG); c.font=F(True,10,CY_FG)
        else:              c.fill=BG(CR_BG); c.font=F(True,10,CR_FG)

        sig = row.get("CS_Signal",""); c = ws.cell(er,CI_SIGNAL,sig)
        c.border=bdr; c.alignment=AL()
        fg,bg = SIGNAL_STYLE.get(sig,("000000",C_WHITE))
        c.font=F(True,9,fg); c.fill=BG(bg)

    for ci,w in {1:9,2:22,3:16,4:8,5:9,
                 6:28,7:13,
                 8:8,9:8,10:8,11:8,12:8,13:7,14:7,15:6,
                 16:6,17:6,18:6,
                 19:6,20:6,21:6,22:6,23:7}.items():
        ws.column_dimensions[CL(ci)].width=w
    for i in range(N_CS): ws.column_dimensions[CL(CS_START+i)].width=9
    ws.column_dimensions[CL(CI_SCORE)].width=7
    ws.column_dimensions[CL(CI_CONV)].width=9
    ws.column_dimensions[CL(CI_SIGNAL)].width=15
    for ri in range(DR, DR+len(df)): ws.row_dimensions[ri].height=16

    for col_i in [9,11,21]:
        rng=f"{CL(col_i)}{DR}:{CL(col_i)}{DR+len(df)-1}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="num",start_value=-0.3,start_color="FFC7CE",
            mid_type="num",mid_value=0,mid_color="FFFFFF",
            end_type="num",end_value=0.5,end_color="C6EFCE"))

    ws.freeze_panes = ws[f"B{DR}"]


def _qc_sheet(wb, df):
    """Quality Compounder sheet — metrics, QC checkmarks, score, signal, EQ badge."""
    ws = wb.create_sheet("Quality Compounder")
    ws.sheet_view.showGridLines = False
    ws.sheet_tab_color = "1A5C2B"
    bdr = BD()

    # Sort by QC_Score desc → 1Y% desc
    sort_cols = [c for c in ["QC_Score", "1Y%"] if c in df.columns]
    df_qc = (df.sort_values(sort_cols, ascending=[False] * len(sort_cols))
               .reset_index(drop=True) if sort_cols else df.reset_index(drop=True))

    DATA_HDRS = [
        ("Ticker",      "Ticker",        9),
        ("Tên Công Ty", "Tên Công Ty",  22),
        ("Sector",      "Sector",        14),
        ("Price ($)",   "Price ($)",      8),
        ("MCap ($B)",   "MCap ($B)",      9),
        ("Moat Score",  "Moat Score",    13),
        ("GM%",         "Gross Margin%",  7),
        ("Op Margin%",  "Op Margin%",     9),
        ("ROIC%",       "ROIC%",          7),
        ("D/E",         "D/E",            6),
        ("FCF/sh",      "FCF/sh",         7),
        ("Curr.R",      "Current Ratio",  7),
        ("EV/EBITDA",   "EV/EBITDA",      9),
        ("P/E",         "P/E",            6),
        ("1Y%",         "1Y%",            7),
    ]
    QC_CHECK_HDRS = [
        ("ROIC✓",  "QC_ROIC"),
        ("OpMgn✓", "QC_OPGM"),
        ("GM✓",    "QC_GM"),
        ("FCF✓",   "QC_FCF"),
        ("DE✓",    "QC_DE"),
        ("Moat✓",  "QC_MOAT"),
    ]
    RESULT_HDRS = [
        ("Score",    "QC_Score"),
        ("Quality",  "QC_Signal"),
        ("EQ Badge", "EQ_Badge"),
    ]

    ND   = len(DATA_HDRS)
    NQC  = len(QC_CHECK_HDRS)
    NR   = len(RESULT_HDRS)
    TOTAL = ND + NQC + NR
    PCT_COLS = {"Gross Margin%", "Op Margin%", "ROIC%", "1Y%"}
    DR = 5

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{CL(TOTAL)}1")
    c = ws["A1"]
    c.value = "📊  QUALITY COMPOUNDER SCREENER"
    c.font = Font(name="Calibri", bold=True, size=15, color="FFFFFF")
    c.fill = BG("0D2137"); c.alignment = AL()
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{CL(TOTAL)}2")
    c = ws["A2"]
    c.value = (f"Source: TradingView + yfinance   |   "
               f"{datetime.now().strftime('%Y-%m-%d %H:%M')}   |   Sorted by QC Score ↓")
    c.font = F(False, 8, "88AACC", True); c.fill = BG("0D2137"); c.alignment = AL()
    ws.row_dimensions[2].height = 16

    # ── Group headers (row 3) ─────────────────────────────────────────────────
    groups = [
        (1,       3,        "IDENTITY",                                    "1C3550"),
        (4,       5,        "PRICE & SIZE",                                "2E5E8E"),
        (6,       6,        "🏰  MOAT",                                    "4A235A"),
        (7,       15,       "📊  QC METRICS",                              "0E3251"),
        (16,      21,       "✅  QC CRITERIA  (✓ = đạt / ✗ = không đạt)", "1A5C2B"),
        (22,      22,       "SCORE",                                       "0D3B1A"),
        (23,      23,       "QUALITY",                                     "0D3B1A"),
        (24,      24,       "EQ BADGE",                                    "0D2020"),
    ]
    ws.row_dimensions[3].height = 14
    for cs, ce, lbl, bg_hex in groups:
        c = ws.cell(3, cs, lbl)
        c.font = F(True, 8, "C8DDEF"); c.fill = BG(bg_hex)
        c.alignment = AL(); c.border = bdr
        if cs < ce:
            ws.merge_cells(f"{CL(cs)}3:{CL(ce)}3")

    # ── Column headers (row 4) ────────────────────────────────────────────────
    ws.row_dimensions[4].height = 28
    for ci, (hdr, _, w) in enumerate(DATA_HDRS, 1):
        c = ws.cell(4, ci, hdr)
        c.font = F(True, 8, C_WHITE); c.fill = BG(C_NAVY)
        c.alignment = AL(wrap=True); c.border = bdr
        ws.column_dimensions[CL(ci)].width = w

    for i, (hdr, _) in enumerate(QC_CHECK_HDRS):
        ci = ND + 1 + i
        c = ws.cell(4, ci, hdr)
        c.font = F(True, 8, "5EC472"); c.fill = BG("0D2A10")
        c.alignment = AL(wrap=True); c.border = bdr
        ws.column_dimensions[CL(ci)].width = 7

    for i, (hdr, _) in enumerate(RESULT_HDRS):
        ci = ND + NQC + 1 + i
        bg_h = "0D3B1A" if i < 2 else "0D2020"
        c = ws.cell(4, ci, hdr)
        c.font = F(True, 8, C_WHITE); c.fill = BG(bg_h)
        c.alignment = AL(wrap=True); c.border = bdr
        ws.column_dimensions[CL(ci)].width = 7 if i == 0 else 16

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(df_qc.iterrows(), 1):
        er  = DR + ri - 1
        alt = ri % 2 == 0
        rbg = C_ALT if alt else C_WHITE
        ws.row_dimensions[er].height = 16

        for ci, (_, key, _) in enumerate(DATA_HDRS, 1):
            val = row.get(key)
            c   = ws.cell(er, ci)
            c.border = bdr; c.fill = BG(rbg)
            c.alignment = AL("left" if ci <= 3 else "center")
            if val is None or (isinstance(val, float) and pd.isna(val)):
                c.value = "—"; c.font = F(size=9, color="BBBBBB"); continue
            if key == "Ticker":
                c.value = val; c.font = F(True, 9, C_NAVY); c.alignment = AL("left")
            elif key == "Tên Công Ty":
                c.value = val; c.font = F(size=8, color="1C3550"); c.alignment = AL("left")
            elif key == "Sector":
                c.value = val; c.font = F(size=8, italic=True, color="334466"); c.alignment = AL("left")
            elif key == "Moat Score":
                fg_m, bg_m = MOAT_SCORE_STYLE.get(val, ("000000", C_WHITE))
                c.value = val; c.font = F(True, 9, fg_m); c.fill = BG(bg_m); c.alignment = AL()
            elif key == "Price ($)":
                c.value = val; c.number_format = '"$"#,##0.00'; c.font = F(size=9)
            elif key == "MCap ($B)":
                c.value = val; c.number_format = "#,##0.0"; c.font = F(size=9)
            elif key in PCT_COLS:
                c.value = val / 100; c.number_format = '+0.0%;(0.0%);"-"'
                c.font = (F(size=9, color=CG_FG, bold=(val > 25)) if val > 0
                          else F(size=9, color=CR_FG) if val < 0 else F(size=9))
            else:
                c.value = val; c.number_format = "0.0"; c.font = F(size=9)

        for i, (_, qc_key) in enumerate(QC_CHECK_HDRS):
            ci = ND + 1 + i
            v  = row.get(qc_key)
            c  = ws.cell(er, ci)
            c.border = bdr; c.alignment = AL()
            if v is True:
                c.value = "✓"; c.font = F(True, 11, CG_FG); c.fill = BG(CG_BG)
            elif v is False:
                c.value = "✗"; c.font = F(True, 10, CR_FG); c.fill = BG(CR_BG)
            else:
                c.value = "–"; c.font = F(size=9, color="AAAAAA"); c.fill = BG("F2F2F2")

        # Score
        sc  = int(row.get("QC_Score", 0) or 0)
        pct = sc / NQC
        c   = ws.cell(er, ND + NQC + 1, sc)
        c.border = bdr; c.alignment = AL(); c.number_format = "0"
        if pct >= 5/6:   c.fill = BG(CG_BG); c.font = F(True, 11, CG_FG)
        elif pct >= 3/6: c.fill = BG(CB_BG); c.font = F(True, 11, CB_FG)
        elif pct >= 1/6: c.fill = BG(CY_BG); c.font = F(True, 11, CY_FG)
        else:            c.fill = BG(CR_BG); c.font = F(True, 11, CR_FG)

        # Quality signal
        sig = row.get("QC_Signal", "")
        fg_s, bg_s = QC_SIGNAL_STYLE.get(sig, ("000000", C_WHITE))
        c = ws.cell(er, ND + NQC + 2, sig or "")
        c.border = bdr; c.alignment = AL()
        c.font = F(True, 9, fg_s); c.fill = BG(bg_s)

        # EQ Badge
        eq = row.get("EQ_Badge")
        c  = ws.cell(er, ND + NQC + 3, eq if eq else "—")
        c.border = bdr; c.alignment = AL()
        if eq and eq in EQ_BADGE_STYLE:
            fg_e, bg_e = EQ_BADGE_STYLE[eq]
            c.font = F(True, 9, fg_e); c.fill = BG(bg_e)
        else:
            c.font = F(size=9, color="AAAAAA"); c.fill = BG(rbg)

    ws.freeze_panes = ws[f"B{DR}"]


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════
def main():
    p=argparse.ArgumentParser()
    p.add_argument("--market", default="america")
    p.add_argument("--top",    type=int, default=300)
    p.add_argument("--output", default=None)
    p.add_argument("--no-yf",  action="store_true",
                   help="Bỏ qua yfinance, dùng TV TTM data cho Moat (nhanh hơn)")
    a=p.parse_args()
    out = a.output or f"tv_top{a.top}_{a.market}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    print(f"\n{'='*60}")
    print(f"  TV SCREENER — TOP {a.top} {a.market.upper()} + CAN SLIM")
    print(f"{'='*60}\n")

    # 1. Lấy data từ TradingView
    df = clean_df(fetch_data(a.market, a.top))

    # 2. Fetch Moat từ yfinance (trừ khi --no-yf)
    moat_cache = None
    if not a.no_yf:
        tickers = df["Ticker"].tolist()
        sectors = dict(zip(df["Ticker"], df["Sector"].fillna("")))
        moat_cache = build_moat_cache(tickers, sectors)
    else:
        print("  ⚡ Bỏ qua yfinance — dùng TV TTM data cho Moat\n")

    # 3. Market direction + Score + Moat
    market_ok, index_1y = fetch_market_direction(a.market)
    df = score_canslim(df, moat_cache, market_ok=market_ok, index_1y=index_1y)

    # Preview
    prev = ["Ticker","Moat Proxy","Moat Score","EPS Qtr%","ROE%","CS_Score","CS_Signal"]
    print(df[[c for c in prev if c in df.columns]].head(10).to_string(index=False))
    print()

    print("  📝 Tạo Excel...", end="", flush=True)
    write_excel(df, out, a.market, a.top)
    print(f" ✅\n  ✨ Done → {out}\n")


if __name__ == "__main__":
    main()